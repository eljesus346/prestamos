from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
from models import Prestamo, Equipo, Usuario
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime, date

router = APIRouter()

# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS COMPARTIDOS
# ══════════════════════════════════════════════════════════════════════════════

def estilo_encabezado(ws, fila, columnas):
    fill = PatternFill("solid", fgColor="1E293B")
    font = Font(color="FFFFFF", bold=True, size=11)
    for col, titulo in enumerate(columnas, 1):
        cell = ws.cell(row=fila, column=col, value=titulo)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


# ── Helpers de estilo para el reporte profesional ─────────────────────────────
COLOR = {
    "header_bg":      "1E3A5F",
    "header_font":    "FFFFFF",
    "title_bg":       "2563EB",
    "title_font":     "FFFFFF",
    "disponible_bg":  "DCFCE7",
    "disponible_fg":  "166534",
    "prestado_bg":    "DBEAFE",
    "prestado_fg":    "1E40AF",
    "atrasado_bg":    "FEF3C7",
    "atrasado_fg":    "92400E",
    "danado_bg":      "FEE2E2",
    "danado_fg":      "991B1B",
    "mant_bg":        "EDE9FE",
    "mant_fg":        "5B21B6",
    "fila_par":       "F8FAFC",
    "fila_impar":     "FFFFFF",
    "border":         "CBD5E1",
}

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(hex_color, bold=False, size=10, name="Arial"):
    return Font(color=hex_color, bold=bold, size=size, name=name)

def _border():
    side = Side(style="thin", color=COLOR["border"])
    return Border(left=side, right=side, top=side, bottom=side)

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def _estado_estilos(estado_calc):
    mapa = {
        "disponible":    (COLOR["disponible_bg"], COLOR["disponible_fg"], "Disponible"),
        "prestado":      (COLOR["prestado_bg"],   COLOR["prestado_fg"],   "Prestado"),
        "atrasado":      (COLOR["atrasado_bg"],   COLOR["atrasado_fg"],   "Atrasado"),
        "dañado":        (COLOR["danado_bg"],      COLOR["danado_fg"],     "Dañado"),
        "mantenimiento": (COLOR["mant_bg"],        COLOR["mant_fg"],       "Mantenimiento"),
    }
    return mapa.get(estado_calc, (COLOR["fila_impar"], "000000", estado_calc))


# ══════════════════════════════════════════════════════════════════════════════
#  EXPORTAR PRÉSTAMOS  (existente sin cambios)
# ══════════════════════════════════════════════════════════════════════════════
@router.get("/prestamos")
def exportar_prestamos(db: Session = Depends(get_db)):
    prestamos = db.query(Prestamo).order_by(Prestamo.id.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Préstamos"

    columnas = ["#", "Usuario", "Documento", "Equipo", "Tipo", "Serial",
                "Fecha Préstamo", "Devolución Est.", "Devolución Real",
                "Estado", "Observaciones"]
    estilo_encabezado(ws, 1, columnas)

    for p in prestamos:
        usuario = db.query(Usuario).filter(Usuario.id == p.id_usuario).first()
        equipo  = db.query(Equipo).filter(Equipo.id  == p.id_equipo).first()
        ws.append([
            p.id,
            usuario.nombre    if usuario else "—",
            usuario.documento if usuario else "—",
            equipo.nombre     if equipo  else "—",
            equipo.tipo       if equipo  else "—",
            equipo.serial     if equipo  else "—",
            p.fecha_prestamo.strftime("%d/%m/%Y %H:%M") if p.fecha_prestamo else "—",
            str(p.fecha_devolucion_estimada) if p.fecha_devolucion_estimada else "—",
            p.fecha_devolucion_real.strftime("%d/%m/%Y %H:%M") if p.fecha_devolucion_real else "—",
            p.estado,
            p.observaciones or "—",
        ])

    anchos = [5, 25, 15, 25, 12, 18, 18, 15, 15, 12, 30]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = ancho

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre = f"prestamos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre}"}
    )


# ══════════════════════════════════════════════════════════════════════════════
#  EXPORTAR EQUIPOS SIMPLE  (existente sin cambios)
# ══════════════════════════════════════════════════════════════════════════════
@router.get("/equipos")
def exportar_equipos(db: Session = Depends(get_db)):
    equipos = db.query(Equipo).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Equipos"

    columnas = ["#", "Nombre", "Tipo", "Serial", "Código Interno",
                "Estado", "Observaciones", "Fecha Registro"]
    estilo_encabezado(ws, 1, columnas)

    for e in equipos:
        ws.append([
            e.id, e.nombre, e.tipo, e.serial,
            e.codigo_interno or "—",
            e.estado,
            e.observaciones or "—",
            e.fecha_registro.strftime("%d/%m/%Y") if e.fecha_registro else "—",
        ])

    anchos = [5, 25, 12, 20, 15, 15, 30, 15]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = ancho

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre = f"equipos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre}"}
    )


# ══════════════════════════════════════════════════════════════════════════════
#  EXPORTAR USUARIOS  (existente sin cambios)
# ══════════════════════════════════════════════════════════════════════════════
@router.get("/usuarios")
def exportar_usuarios(db: Session = Depends(get_db)):
    usuarios = db.query(Usuario).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    columnas = ["#", "Nombre", "Documento", "Teléfono",
                "Correo", "Cargo", "Área", "Estado", "Fecha Registro"]
    estilo_encabezado(ws, 1, columnas)

    for u in usuarios:
        ws.append([
            u.id, u.nombre, u.documento,
            u.telefono or "—", u.correo or "—",
            u.cargo or "—", u.area or "—",
            "Activo" if u.estado else "Inactivo",
            u.fecha_registro.strftime("%d/%m/%Y") if u.fecha_registro else "—",
        ])

    anchos = [5, 25, 15, 15, 28, 20, 20, 10, 15]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = ancho

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre = f"usuarios_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre}"}
    )


# ══════════════════════════════════════════════════════════════════════════════
#  REPORTE DIARIO DE EQUIPOS  ← NUEVO
# ══════════════════════════════════════════════════════════════════════════════
@router.get("/equipos/reporte-diario")
def reporte_diario_equipos(db: Session = Depends(get_db)):
    hoy     = date.today()
    equipos = db.query(Equipo).order_by(Equipo.id.asc()).all()

    # Préstamos activos indexados por id_equipo
    prestamos_activos: dict[int, tuple] = {}
    for p in db.query(Prestamo).filter(Prestamo.estado == "prestado").all():
        usuario = db.query(Usuario).filter(Usuario.id == p.id_usuario).first()
        prestamos_activos[p.id_equipo] = (p, usuario)

    # Contadores previos (para barra de resumen)
    contadores = {"disponible": 0, "prestado": 0, "atrasado": 0,
                  "dañado": 0, "mantenimiento": 0}
    for e in equipos:
        est = e.estado
        if est == "prestado" and e.id in prestamos_activos:
            p, _ = prestamos_activos[e.id]
            if p.fecha_devolucion_estimada and p.fecha_devolucion_estimada < hoy:
                est = "atrasado"
        contadores[est] = contadores.get(est, 0) + 1

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    #  HOJA 1: Reporte completo de todos los equipos
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Reporte de Equipos"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"

    # Título
    ws.merge_cells("A1:M1")
    c = ws["A1"]
    c.value = "REPORTE DIARIO DE EQUIPOS"
    c.fill  = _fill(COLOR["title_bg"])
    c.font  = _font(COLOR["title_font"], bold=True, size=14)
    c.alignment = _center()
    ws.row_dimensions[1].height = 34

    # Subtítulo
    ws.merge_cells("A2:M2")
    c = ws["A2"]
    c.value = (f"Generado el {datetime.now().strftime('%d/%m/%Y')} "
               f"a las {datetime.now().strftime('%H:%M')}  |  "
               f"Sistema de Préstamos de Equipos")
    c.fill  = _fill("EFF6FF")
    c.font  = _font("475569", size=9)
    c.alignment = _center()
    ws.row_dimensions[2].height = 18

    # Separador
    ws.merge_cells("A3:M3")
    ws["A3"].fill = _fill("E2E8F0")
    ws.row_dimensions[3].height = 6

    # Barra de resumen (fila 4)
    ws.row_dimensions[4].height = 28
    bloques = [
        ("A4", "B4", "Disponibles",   contadores["disponible"],   COLOR["disponible_bg"], COLOR["disponible_fg"]),
        ("C4", "D4", "Prestados",     contadores["prestado"],     COLOR["prestado_bg"],   COLOR["prestado_fg"]),
        ("E4", "F4", "Atrasados",     contadores["atrasado"],     COLOR["atrasado_bg"],   COLOR["atrasado_fg"]),
        ("G4", "H4", "Dañados",       contadores["dañado"],       COLOR["danado_bg"],     COLOR["danado_fg"]),
        ("I4", "J4", "Mantenimiento", contadores["mantenimiento"],COLOR["mant_bg"],       COLOR["mant_fg"]),
        ("K4", "M4", "Total equipos", len(equipos),               "F1F5F9",               "334155"),
    ]
    for c1, c2, label, valor, bg, fg in bloques:
        ws.merge_cells(f"{c1}:{c2}")
        cell = ws[c1]
        cell.value      = f"{label}:  {valor}"
        cell.fill       = _fill(bg)
        cell.font       = _font(fg, bold=True, size=10)
        cell.alignment  = _center()
        cell.border     = _border()

    # Separador
    ws.merge_cells("A5:M5")
    ws["A5"].fill = _fill("E2E8F0")
    ws.row_dimensions[5].height = 6

    # Leyenda de colores (fila 6)
    ws.row_dimensions[6].height = 16
    leyenda = [
        ("A6", "B6", "Verde = Disponible",    COLOR["disponible_bg"], COLOR["disponible_fg"]),
        ("C6", "D6", "Azul = Prestado",       COLOR["prestado_bg"],   COLOR["prestado_fg"]),
        ("E6", "F6", "Naranja = Atrasado",    COLOR["atrasado_bg"],   COLOR["atrasado_fg"]),
        ("G6", "H6", "Rojo = Dañado",         COLOR["danado_bg"],     COLOR["danado_fg"]),
        ("I6", "J6", "Morado = Mantenimiento",COLOR["mant_bg"],       COLOR["mant_fg"]),
    ]
    for c1, c2, label, bg, fg in leyenda:
        ws.merge_cells(f"{c1}:{c2}")
        cell = ws[c1]
        cell.value     = label
        cell.fill      = _fill(bg)
        cell.font      = _font(fg, size=8)
        cell.alignment = _center()
        cell.border    = _border()

    # Encabezados de tabla (fila 7)
    columnas_hoja1 = [
        ("#",               5),
        ("Nombre del Equipo", 28),
        ("Tipo",            12),
        ("Serial",          20),
        ("Cód. Interno",    14),
        ("Estado",          16),
        ("Responsable",     26),
        ("Documento",       14),
        ("Fecha Préstamo",  17),
        ("Dev. Estimada",   14),
        ("Días Atraso",     12),
        ("Teléfono",        14),
        ("Observaciones",   32),
    ]
    ws.row_dimensions[7].height = 24
    for ci, (titulo, ancho) in enumerate(columnas_hoja1, 1):
        cell = ws.cell(row=7, column=ci, value=titulo)
        cell.fill      = _fill(COLOR["header_bg"])
        cell.font      = _font(COLOR["header_font"], bold=True, size=10)
        cell.alignment = _center()
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = ancho

    # Filas de datos
    for idx, e in enumerate(equipos):
        row = 8 + idx
        ws.row_dimensions[row].height = 20

        estado_calc = e.estado
        pi          = prestamos_activos.get(e.id)
        dias_atraso = 0

        if e.estado == "prestado" and pi:
            p, u = pi
            if p.fecha_devolucion_estimada and p.fecha_devolucion_estimada < hoy:
                estado_calc = "atrasado"
                dias_atraso = (hoy - p.fecha_devolucion_estimada).days

        bg, fg, etiqueta = _estado_estilos(estado_calc)

        # Color de fila completa
        if estado_calc in ("atrasado", "dañado", "mantenimiento"):
            fila_bg = bg          # Toda la fila con el color del estado
        elif estado_calc == "prestado":
            fila_bg = "F0F7FF"    # Azul muy suave para prestados normales
        else:
            fila_bg = COLOR["fila_par"] if idx % 2 == 0 else COLOR["fila_impar"]

        # Datos de préstamo
        responsable = doc_res = tel_res = fecha_p = fecha_d = dias_txt = "—"
        if pi and e.estado == "prestado":
            p, u = pi
            if u:
                responsable = u.nombre
                doc_res     = u.documento
                tel_res     = u.telefono or "—"
            fecha_p = p.fecha_prestamo.strftime("%d/%m/%Y %H:%M") if p.fecha_prestamo else "—"
            fecha_d = p.fecha_devolucion_estimada.strftime("%d/%m/%Y") if p.fecha_devolucion_estimada else "—"
            if dias_atraso > 0:
                dias_txt = f"{dias_atraso} día{'s' if dias_atraso > 1 else ''}"

        valores = [
            e.id, e.nombre, e.tipo.capitalize(), e.serial,
            e.codigo_interno or "—",
            etiqueta,
            responsable, doc_res, fecha_p, fecha_d, dias_txt,
            tel_res,
            e.observaciones or "—",
        ]
        alineaciones = ["center","left","center","center","center",
                        "center","left","center","center","center",
                        "center","center","left"]

        for ci, (valor, align) in enumerate(zip(valores, alineaciones), 1):
            cell = ws.cell(row=row, column=ci, value=valor)
            cell.border    = _border()
            cell.alignment = Alignment(horizontal=align, vertical="center",
                                       wrap_text=(ci == 13))
            if ci == 6:                             # Columna estado
                cell.fill = _fill(bg)
                cell.font = _font(fg, bold=True, size=9)
            elif ci == 11 and dias_atraso > 0:      # Columna días atraso
                cell.fill = _fill(COLOR["atrasado_bg"])
                cell.font = _font(COLOR["atrasado_fg"], bold=True, size=9)
            else:
                cell.fill = _fill(fila_bg)
                cell.font = _font("374151", size=9)

    # ══════════════════════════════════════════════════════════════════════════
    #  HOJA 2: Solo préstamos activos y atrasados
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Prestamos Activos")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A4"

    ws2.merge_cells("A1:J1")
    c = ws2["A1"]
    c.value = "PRESTAMOS ACTIVOS  —  Equipos en manos de colaboradores"
    c.fill  = _fill(COLOR["title_bg"])
    c.font  = _font(COLOR["title_font"], bold=True, size=13)
    c.alignment = _center()
    ws2.row_dimensions[1].height = 30

    ws2.merge_cells("A2:J2")
    c = ws2["A2"]
    c.value = (f"Fecha de corte: {hoy.strftime('%d/%m/%Y')}  |  "
               f"Atrasados: {contadores['atrasado']}  |  "
               f"Prestados (en tiempo): {contadores['prestado']}")
    c.fill  = _fill("EFF6FF")
    c.font  = _font("475569", size=9)
    c.alignment = _center()
    ws2.row_dimensions[2].height = 18

    cols2 = [
        ("#",              5),  ("Equipo",         26), ("Tipo",          12),
        ("Serial",        20),  ("Estado",         16), ("Responsable",   26),
        ("Documento",     14),  ("Fecha Préstamo", 17), ("Dev. Estimada", 14),
        ("Días Atraso",   12),
    ]
    ws2.row_dimensions[3].height = 24
    for ci, (titulo, ancho) in enumerate(cols2, 1):
        cell = ws2.cell(row=3, column=ci, value=titulo)
        cell.fill      = _fill(COLOR["header_bg"])
        cell.font      = _font(COLOR["header_font"], bold=True, size=10)
        cell.alignment = _center()
        cell.border    = _border()
        ws2.column_dimensions[get_column_letter(ci)].width = ancho

    fila2 = 4
    total_activos = 0
    for idx2, e in enumerate(equipos):
        if e.estado != "prestado":
            continue
        pi = prestamos_activos.get(e.id)
        if not pi:
            continue
        p, u = pi

        estado_calc2 = "prestado"
        dias2        = 0
        if p.fecha_devolucion_estimada and p.fecha_devolucion_estimada < hoy:
            estado_calc2 = "atrasado"
            dias2        = (hoy - p.fecha_devolucion_estimada).days

        bg2, fg2, etiqueta2 = _estado_estilos(estado_calc2)
        fila_bg2 = bg2 if estado_calc2 == "atrasado" else (
            COLOR["fila_par"] if total_activos % 2 == 0 else COLOR["fila_impar"]
        )

        ws2.row_dimensions[fila2].height = 20
        fecha_p2 = p.fecha_prestamo.strftime("%d/%m/%Y %H:%M") if p.fecha_prestamo else "—"
        fecha_d2 = p.fecha_devolucion_estimada.strftime("%d/%m/%Y") if p.fecha_devolucion_estimada else "—"
        dias_txt2 = f"{dias2} día{'s' if dias2 > 1 else ''}" if dias2 > 0 else "—"

        valores2 = [
            e.id, e.nombre, e.tipo.capitalize(), e.serial,
            etiqueta2,
            u.nombre    if u else "—",
            u.documento if u else "—",
            fecha_p2, fecha_d2, dias_txt2,
        ]
        for ci, val in enumerate(valores2, 1):
            cell = ws2.cell(row=fila2, column=ci, value=val)
            cell.border    = _border()
            cell.alignment = _left() if ci == 6 else _center()
            if ci == 5:
                cell.fill = _fill(bg2);  cell.font = _font(fg2, bold=True, size=9)
            elif ci == 10 and dias2 > 0:
                cell.fill = _fill(COLOR["atrasado_bg"])
                cell.font = _font(COLOR["atrasado_fg"], bold=True, size=9)
            else:
                cell.fill = _fill(fila_bg2)
                cell.font = _font("374151", size=9)

        fila2 += 1
        total_activos += 1

    if total_activos == 0:
        ws2.merge_cells("A4:J4")
        c = ws2["A4"]
        c.value      = "No hay préstamos activos en este momento."
        c.font       = _font("64748B", size=10)
        c.alignment  = _center()
        c.fill       = _fill("F8FAFC")

    # ── Guardar ───────────────────────────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre = f"reporte_equipos_{hoy.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre}"}
    )